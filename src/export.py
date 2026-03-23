import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def export_to_csv(cbo_df: pd.DataFrame, filepath: str | Path) -> None:
    depara = cbo_df[['CODIGO', 'TITULO', 'COD_PNAD', 
                     'GRANDE_GRUPO_COD', 'GRANDE_GRUPO_NOME', 
                     'SUBGRUPO_PRINCIPAL_COD', 'SUBGRUPO_COD']].copy()
    
    depara.columns = ['CBO_2002_6dig', 'TITULO_CBO', 'COD_PNAD_4dig',
                      'GG_COD', 'GG_NOME', 'SUBGRUPO_PRINCIPAL', 'SUBGRUPO']
    
    depara.to_csv(filepath, index=False, encoding='utf-8-sig', sep=';')

def _write_sheet(ws, title_text, note_text, headers, rows, col_widths):
    H_FILL = PatternFill('solid', fgColor='1F4E79')
    S_FILL = PatternFill('solid', fgColor='2E75B6')
    A_FILL = PatternFill('solid', fgColor='D6E4F0')
    W_FILL = PatternFill('solid', fgColor='FFFFFF')
    N_FILL = PatternFill('solid', fgColor='EBF3FB')
    H_FONT = Font(bold=True, color='FFFFFF', size=10)
    N_FONT = Font(size=9)

    n_cols = len(headers)
    last_col = get_column_letter(n_cols)

    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = title_text
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=13)
    ws['A1'].fill = H_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = note_text
    ws['A2'].font = Font(italic=True, size=9)
    ws['A2'].fill = N_FILL
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[2].height = 30

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = H_FONT; c.fill = S_FILL
        c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
    ws.row_dimensions[3].height = 36

    for ri, row in enumerate(rows, 4):
        fill = A_FILL if ri % 2 == 0 else W_FILL
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = N_FONT; c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical='top')

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{last_col}{ws.max_row}'

def export_to_excel(cbo_df: pd.DataFrame, correspondencia_df: pd.DataFrame, filepath: str | Path) -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'COD-PNAD x CBO'
    
    headers_ws1 = correspondencia_df.columns.tolist()
    widths_ws1 = [20] * len(headers_ws1)

    _write_sheet(
        ws1,
        'Correspondência COD-PNAD (4 dígitos) ↔ CBO-2002 (6 dígitos)',
        'Regra: COD-PNAD = 4 primeiros dígitos do CBO-2002.',
        headers_ws1,
        correspondencia_df.values.tolist(),
        widths_ws1
    )

    ws2 = wb.create_sheet('CBO-2002 Completa')
    _write_sheet(
        ws2,
        'CBO-2002 Completa com Código COD-PNAD',
        'Busque pelo CODIGO_CBO_6dig e retorne COD_PNAD_4dig.',
        ['CODIGO CBO-2002 (6 dig)', 'TITULO OCUPAÇÃO', 'COD-PNAD (4 dig)',
         'GRANDE GRUPO NOME', 'GG COD', 'SUBGRUPO PRINC.', 'SUBGRUPO'],
        cbo_df[['CODIGO', 'TITULO', 'COD_PNAD',
                'GRANDE_GRUPO_NOME', 'GRANDE_GRUPO_COD',
                'SUBGRUPO_PRINCIPAL_COD', 'SUBGRUPO_COD']].values.tolist(),
        [22, 52, 16, 50, 10, 20, 16]
    )

    wb.save(filepath)